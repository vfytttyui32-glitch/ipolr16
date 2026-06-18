from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Product, Category, Manufacturer, Cart, CartItem

def product_list(request):
    products = Product.objects.all()
    
    search_query = request.GET.get('search', '').strip()
    category_id = request.GET.get('category', '')
    manufacturer_id = request.GET.get('manufacturer', '')
    
    # Все фильтры применяются вместе через Q
    filters = Q()
    
    if search_query:
        filters &= Q(name__icontains=search_query) | Q(description__icontains=search_query)
    
    if category_id:
        filters &= Q(category_id=category_id)
    
    if manufacturer_id:
        filters &= Q(manufacturer_id=manufacturer_id)
    
    products = products.filter(filters)
    
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()
    
    context = {
        'products': products,
        'categories': categories,
        'manufacturers': manufacturers,
        'selected_category': int(category_id) if category_id else None,
        'selected_manufacturer': int(manufacturer_id) if manufacturer_id else None,
        'search_query': search_query,
    }
    return render(request, 'product_list.html', context)

def product_detail(request, pk):
    """Детальная информация о товаре"""
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'product_detail.html', {'product': product})

@login_required
def add_to_cart(request, product_id):
    """Добавление товара в корзину"""
    product = get_object_or_404(Product, pk=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    
    cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)
    if not created:
        cart_item.quantity += 1
        cart_item.save()
    
    return redirect('cart_view')

@login_required
def update_cart(request, item_id):
    """Обновление количества товара в корзине"""
    cart_item = get_object_or_404(CartItem, pk=item_id)
    
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        # Валидация: не больше, чем в наличии
        if quantity <= cart_item.product.quantity_in_stock:
            cart_item.quantity = quantity
            cart_item.save()
    
    return redirect('cart_view')

@login_required
def remove_from_cart(request, item_id):
    """Удаление товара из корзины"""
    cart_item = get_object_or_404(CartItem, pk=item_id)
    cart_item.delete()
    return redirect('cart_view')

@login_required
def cart_view(request):
    """Просмотр корзины"""
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.cartitem_set.all()
    
    total_price = sum(item.get_total_price() for item in cart_items)
    
    return render(request, 'cart.html', {
        'cart_items': cart_items,
        'total_price': total_price,
    })

import openpyxl
from openpyxl.styles import Font
from django.core.mail import EmailMessage
from django.http import HttpResponse
import io

@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.cartitem_set.select_related('product').all()

    if not cart_items.exists():
        return redirect('cart_view')

    if request.method == 'POST':
        address = request.POST.get('address', '')

        # Генерация Excel чека
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Чек'

        # Заголовок
        ws['A1'] = 'Чек заказа'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Покупатель: {request.user.email}'
        ws['A3'] = f'Адрес доставки: {address}'
        ws['A4'] = ''

        # Шапка таблицы
        headers = ['№', 'Товар', 'Цена', 'Количество', 'Сумма']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)

        # Товары
        total = 0
        for i, item in enumerate(cart_items, 1):
            item_total = item.get_total_price()
            total += item_total
            ws.cell(row=5 + i, column=1, value=i)
            ws.cell(row=5 + i, column=2, value=item.product.name)
            ws.cell(row=5 + i, column=3, value=float(item.product.price))
            ws.cell(row=5 + i, column=4, value=item.quantity)
            ws.cell(row=5 + i, column=5, value=float(item_total))

        # Итого
        total_row = 5 + len(cart_items) + 1
        ws.cell(row=total_row, column=4, value='Итого:').font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(total)).font = Font(bold=True)

        # Ширина колонок
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Сохранить в память
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Отправка email
        email = EmailMessage(
            subject='Ваш заказ оформлен',
            body=f'Здравствуйте, {request.user.username}!\n\nВаш заказ оформлен. Чек во вложении.\n\nАдрес доставки: {address}\nСумма заказа: {total} руб.',
            to=[request.user.email],
        )
        email.attach('check.xlsx', buffer.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()

        # Очистка корзины
        cart_items.delete()

        return render(request, 'checkout.html', {'success': True, 'total': total})

    total_price = sum(item.get_total_price() for item in cart_items)
    return render(request, 'checkout.html', {
        'cart_items': cart_items,
        'total_price': total_price,
    })
try:
    email.send()
    print("Письмо отправлено успешно!")
except Exception as e:
    print(f"Ошибка отправки: {e}")




